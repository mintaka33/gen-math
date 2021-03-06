from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from PyPDF4.pdf import PdfFileReader, PdfFileWriter
from datetime import datetime
from win32com import client
import win32api
import numpy as np
import random
import sys
import os

num_page = 10
min1, max1 = 15, 40
min2, max2 = 15, 40

row_per_page = 25
col_per_page = 2
math_group_cols = 6
tests_per_page = row_per_page * col_per_page
math_group_rows = row_per_page * num_page
math_group_num = col_per_page

def gen_outfile_name():
    n = datetime.now()
    time_str = 'math_'
    time_str += str(n.year) + '-' + str(n.month).zfill(2) + '-' + str(n.day).zfill(2)
    time_str += '_' + str(n.hour).zfill(2) + '-' + str(n.minute).zfill(2) + '-' + str(n.second).zfill(2)
    time_str += '_' + str(n.microsecond).zfill(6)
    time_str += '.xlsx'
    return time_str

class MathTests():
    def __init__(self):
        self.testset_add = []
        self.testset_sub = []
        self.testset_add_flex = []
        self.testset_sub_flex = []
        for i in range(min1, max1):
            for j in range(min2, max2):
                test = [str(i), '+', str(j), '=', ' ', ' ']
                self.testset_add.append(test)
        print('INFO: total', len(self.testset_add), 'tests in ADD set')
        for i in range(min1, max1):
            for j in range(min2, max2):
                test1 = [str(i), '+', ' ', '=', str(i+j), ' ']
                test2 = [' ', '+', str(j), '=', str(i+j), ' ']
                self.testset_add_flex.append(test1)
                self.testset_add_flex.append(test2)
        print('INFO: total', len(self.testset_add_flex), 'tests in ADD_Flex set')
        for i in range(min1, max1):
            for j in range(min2, max2):
                if (i - j) > 0:
                    test = [str(i), '-', str(j), '=', ' ', ' ']
                    self.testset_sub.append(test)
        print('INFO: total', len(self.testset_sub), 'tests in SUB set')
        for i in range(min1, max1):
            for j in range(min2, max2):
                if (i - j) > 0:
                    test1 = [str(i), '-', ' ', '=', str(i-j), ' ']
                    test2 = [' ', '-', str(j), '=', str(i-j), ' ']
                    self.testset_sub_flex.append(test1)
                    self.testset_sub_flex.append(test2)
        print('INFO: total', len(self.testset_sub_flex), 'tests in SUB_Flex set')

    def get_tests_add(self, num):
        tests = []
        test_array = np.random.randint(len(self.testset_add), size=(num))
        for i in range(num):
            tests.append(self.testset_add[test_array[i]])
        return tests

    def get_tests_add_flex(self, num):
        tests = []
        test_array = np.random.randint(len(self.testset_add_flex), size=(num))
        for i in range(num):
            tests.append(self.testset_add_flex[test_array[i]])
        return tests

    def get_tests_sub(self, num):
        tests = []
        test_array = np.random.randint(len(self.testset_sub), size=(num))
        for i in range(num):
            tests.append(self.testset_sub[test_array[i]])
        return tests

    def get_tests_sub_flex(self, num):
        tests = []
        test_array = np.random.randint(len(self.testset_sub_flex), size=(num))
        for i in range(num):
            tests.append(self.testset_sub_flex[test_array[i]])
        return tests

    def get_tests_mix(self, num):
        add_tests = self.get_tests_add(int(num/2))
        sub_tests = self.get_tests_sub(int(num/2))
        mix_tests = add_tests + sub_tests
        random.shuffle(mix_tests)
        return mix_tests

    def get_tests_flex_mix(self, num):
        add_flex_tests = self.get_tests_add_flex(int(num/2))
        sub_flex_tests = self.get_tests_sub_flex(int(num/2))
        mix_flex_tests = add_flex_tests + sub_flex_tests
        random.shuffle(mix_flex_tests)
        return mix_flex_tests

class XmlWriter():
    def __init__(self):
        self.dest_filename = gen_outfile_name()
        self.total_cols = math_group_cols * math_group_num
        self.total_rows = math_group_rows
        self.font_size = 21
        self.column_width = 7
        self.mystyle = NamedStyle(name="mystyle")
        self.mystyle.font = Font(name='Arial', size=self.font_size)
        self.mystyle.alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
        self.wb = Workbook()
        self.wb.add_named_style(self.mystyle)
        self.ws = self.wb.active
        self.ws.title = "math"
        self.ws.page_margins.left = 0.5
        self.ws.page_margins.right = 0.5
        self.ws.page_margins.top = 0.75
        self.ws.page_margins.bottom = 0.75
        self.ws.page_setup.fitToWidth = 1
        self.ws.print_options.horizontalCentered = True
        self.ws.print_options.verticalCentered = True
        self.ws.print_options.gridLines = True
        # set column width
        for col_idx in range(self.total_cols):
            self.ws.column_dimensions[get_column_letter(col_idx+1)].width = self.column_width

    def write_group(self, group_idx, tests):
        print('INFO: start to write excel column with %d tests...' % len(tests))
        start = group_idx * math_group_cols + 1
        end = start + math_group_cols - 1
        for row_idx, row in enumerate(self.ws.iter_rows(min_col=start, max_row=len(tests), max_col=end)):
            test = tests[row_idx]
            for col_idx, cell in enumerate(row):
                cell.style = self.mystyle
                cell.value = test[col_idx]
        print('INFO: done. write one excel column')

    def export_to_pdf(self):
        print('INFO: start to export xlsx as PDF...')
        xlsx_file = self.dest_filename
        pdf_file = xlsx_file.split('.xlsx')[0] + '.pdf'
        full_infile = os.getcwd() + '\\' + xlsx_file
        full_outfile = os.getcwd() + '\\' + pdf_file
        app = client.DispatchEx("Excel.Application")
        app.Interactive = False
        app.Visible = False
        Workbook = app.Workbooks.Open(full_infile)
        try:
            Workbook.ActiveSheet.ExportAsFixedFormat(0, full_outfile)
            print('INFO: done. export PDF file to', full_outfile)
        except Exception as e:
            print("ERROR: failed to convert in PDF format")
            print(str(e))
        finally:
            print('INFO: export app exit')
            Workbook.Close()
            app.Quit()
            os.system('del ' + full_infile)
            return pdf_file

    def save_to_file(self):
        print('INFO: start to save excel xlsx to file...')
        self.wb.save(self.dest_filename)
        print('INFO: done. write xlsx file to', self.dest_filename)
        return self.export_to_pdf()

def gen_math_tests(test_types):
    math = MathTests()
    row0, row1 = [], []
    for test_type in test_types:
        if test_type == 'add':
            row0 += math.get_tests_add(math_group_rows)
            row1 += math.get_tests_add(math_group_rows)
        elif test_type == 'sub':
            row0 += math.get_tests_sub(math_group_rows)
            row1 += math.get_tests_sub(math_group_rows)
        elif test_type == 'add_sub':
            row0 += math.get_tests_add(math_group_rows)
            row1 += math.get_tests_sub(math_group_rows)
        elif test_type == 'mix':
            row0 += math.get_tests_mix(math_group_rows)
            row1 += math.get_tests_mix(math_group_rows)
        elif test_type == 'flex':
            row0 += math.get_tests_add_flex(math_group_rows)
            row1 += math.get_tests_sub_flex(math_group_rows)
        elif test_type == 'flex_mix':
            row0 += math.get_tests_flex_mix(math_group_rows)
            row1 += math.get_tests_flex_mix(math_group_rows)
    test_num = len(row0)+len(row1)
    page_num = (test_num + tests_per_page - 1)/ (tests_per_page)
    print('INFO: %d tests in %d pages generated' %(test_num, page_num))
    return row0, row1

def export_to_pdf(row0, row1):
    xml = XmlWriter()
    if len(row0) != len(row1) or len(row0) < 1:
        print('ERROR: invalid math tests')
        return
    # write tests to xml file
    xml.write_group(0, row0)
    xml.write_group(1, row1)
    # save xml file
    return xml.save_to_file()

def createPagePdf(num, tmp):
    c = canvas.Canvas(tmp)
    for i in range(1, num + 1):
        c.setFontSize(24)
        c.drawString((210 // 2) * mm, (8) * mm, str(i))
        c.showPage()
    c.save()

def add_page_numgers(pdf_path):
    tmp = "__tmp.pdf"
    output = PdfFileWriter()
    with open(pdf_path, 'rb') as f:
        pdf = PdfFileReader(f, strict=False)
        n = pdf.getNumPages()
        print('INFO: PDF file number = %d' % n)
        # create new PDF with page numbers
        print('INFO: Start to create page PDF file')
        createPagePdf(n, tmp)
        print('INFO: Page PDF file created, start to merge with original PDF...')
        with open(tmp, 'rb') as ftmp:
            numberPdf = PdfFileReader(ftmp)
            # iterarte pages
            for p in range(n):
                page = pdf.getPage(p)
                numberLayer = numberPdf.getPage(p)
                # merge number page with actual page
                page.mergePage(numberLayer)
                output.addPage(page)
            # write result
            if output.getNumPages():
                newpath = pdf_path[:-4] + "_num.pdf"
                with open(newpath, 'wb') as f:
                    output.write(f)
        os.remove(tmp)
        return newpath

def helper():
    print("ERROR: invalid command line. example: python math.py ['add', 'sub', 'add_sub', 'mix']")
    print(" Valid test_type: 'add', 'sub', 'add_sub', 'mix', 'flex', 'flex_mix' ")
    exit()

if __name__ == "__main__":
    r0, r1 = gen_math_tests(['add', 'sub', 'add_sub', 'mix', 'flex', 'flex_mix'])
    pdf_filename = export_to_pdf(r0, r1)
    print('INFO: Generated PDF file %s'%pdf_filename)
    print('INFO: Start to add PDF page number...')
    new_pdfname = add_page_numgers(pdf_filename)
    print('INFO: Generated PDF file %s with page number '%new_pdfname)
    os.remove(pdf_filename)
    print('done')
